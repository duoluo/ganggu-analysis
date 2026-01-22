#!/usr/bin/env python3
"""
港股打新收益分析系统 - 数据生成脚本
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
from datetime import datetime

# 分成比例配置
COMMISSION_RATES = {
    '0%': ['王峰', '安粉燕', '安香叶', '安海燕', '王香叶', '安三娃'],
    '30%': ['刘彬', '芦云', '郜燕燕', '姚智鑫', '李焱', '叶通', '赵苗苗', '芦川', '杨凯'],
    '50%': ['李文博', '李妍妍', '郑爽爽', '刘增运(姐夫)', '张连梅(妈妈)', '李明华(大姐)', '李维峰(李爸爸)'],
    '35%': []  # 将在后面自动填充
}

# 管理账户分组
ACCOUNT_GROUPS = {
    '文博管理账户': [
        '李文博', '李妍妍', '郑爽爽', '刘增运(姐夫)', '张连梅(妈妈)', '李明华(大姐)', '李维峰(李爸爸)',
        '张芳针(姨)', '张继前(舅)', '张卓祥（哥）', '唐莎（嫂子）', '张年光（哥）', '茹雅倩（嫂子）',
        '张静', '黄振博', '刘家乐', '刘尊博', '吴清汪', '周立良', '申嵩', '张慧博', '郑路遥',
        '史澳博', '高宁宁', '郑加豪', '万贲', '王敏卓', '高翔', '杨森', '许经菊', '黄秋月',
        '宋亚祥', '王伟', '李芝灵(妗子)'
    ],
    '王峰管理账户': ['刘彬', '芦云', '郜燕燕', '姚智鑫', '李焱', '叶通', '赵苗苗', '芦川', '杨凯'],
    '无分成账户': ['王峰', '安粉燕', '安香叶', '安海燕', '王香叶', '安三娃']
}

# 特殊分成规则（固定金额）
SPECIAL_COMMISSIONS = {
    ('李焱', '蓝思科技'): 300,
    ('姚智鑫', '维立志博'): 300,
    ('姚智鑫', '中慧生物'): 300,
    ('刘彬', '维立志博'): 300
}

# 特定范围分成统计的账户列表
SPECIAL_RANGE_ACCOUNTS = [
    '李文博', '李妍妍', '郑爽爽', '刘增运(姐夫)', '张连梅(妈妈)', '李明华(大姐)', '李维峰(李爸爸)',
    '张芳针(姨)', '张继前(舅)', '张卓祥（哥）', '唐莎（嫂子）', '张年光（哥）', '茹雅倩（嫂子）',
    '张静', '黄振博', '刘家乐', '刘尊博', '吴清汪', '周立良', '申嵩', '张慧博', '郑路遥',
    '史澳博', '高宁宁', '郑加豪', '万贲', '王敏卓', '高翔', '杨森', '许经菊', '黄秋月',
    '宋亚祥', '王伟', '李芝灵(妗子)'
]

# 需要额外加紫金国际分成的账户
ZIJIN_EXTRA_ACCOUNTS = ['郑爽爽', '李芝灵(妗子)', '王敏卓']

print("=" * 100)
print("港股打新收益分析系统 - 数据生成")
print("=" * 100)

# 加载工作簿
print("\n[1/6] 加载Excel文件...")
wb = load_workbook('/Users/wright/Documents/WFCode/GangGu/2025收益计算.xlsx', data_only=True)
ws = wb.active

# 获取标题行
row1 = list(ws[1])  # 新股名称
row2 = list(ws[2])  # 列标题

# 识别所有新股
print("[2/6] 识别所有新股...")
stocks = []
current_stock = None
start_col = None

for i in range(len(row1)):
    if row1[i].value and i > 0:  # 跳过第一列（账户名称）
        if current_stock:
            stocks.append({
                'name': current_stock,
                'start_col': start_col,
                'end_col': i - 1
            })
        current_stock = row1[i].value
        start_col = i

# 添加最后一个股票
if current_stock:
    stocks.append({
        'name': current_stock,
        'start_col': start_col,
        'end_col': len(row1) - 1
    })

# 为每个股票找到收益列
for stock in stocks:
    for i in range(stock['start_col'], stock['end_col'] + 1):
        col_title = str(row2[i].value) if row2[i].value else ""
        if "收益" in col_title:
            stock['shouyi_col'] = i
            break

print(f"   找到 {len(stocks)} 个新股")

# 获取所有账户
print("[3/6] 获取所有账户...")
accounts = []
account_rows = {}  # 账户名 -> 行号

for row_idx in range(4, ws.max_row + 1):
    account_name = ws.cell(row=row_idx, column=1).value
    if account_name and account_name not in ["所有账户总收益", "收益总计", "中签数量总计"]:
        accounts.append(account_name)
        account_rows[account_name] = row_idx

print(f"   找到 {len(accounts)} 个账户")

# 自动填充35%组（所有不在0%、30%和50%组的账户）
all_0_30_50 = COMMISSION_RATES['0%'] + COMMISSION_RATES['30%'] + COMMISSION_RATES['50%']
COMMISSION_RATES['35%'] = [acc for acc in accounts if acc not in all_0_30_50]

print(f"   分成比例分组:")
print(f"     0%组: {len(COMMISSION_RATES['0%'])} 个账户")
print(f"     30%组: {len(COMMISSION_RATES['30%'])} 个账户")
print(f"     50%组: {len(COMMISSION_RATES['50%'])} 个账户")
print(f"     35%组: {len(COMMISSION_RATES['35%'])} 个账户")

# 计算每个股票的总收益
print("[4/6] 计算股票收益...")
stock_revenues = []

for stock in stocks:
    if 'shouyi_col' not in stock:
        continue

    total = 0
    shouyi_col = stock['shouyi_col']

    for account in accounts:
        row_idx = account_rows[account]
        value = ws.cell(row=row_idx, column=shouyi_col + 1).value
        if value is not None and isinstance(value, (int, float)):
            total += value

    stock_revenues.append({
        'name': stock['name'],
        'revenue': round(total, 2),
        'col_index': shouyi_col
    })

print(f"   计算完成，总收益: {sum(s['revenue'] for s in stock_revenues):,.2f}")

# 计算每个账户的收益详情
print("[5/6] 计算账户收益和分成...")
account_details = []

for account in accounts:
    row_idx = account_rows[account]

    # 确定分成比例
    if account in COMMISSION_RATES['0%']:
        rate = 0.0
        rate_group = '0%'
    elif account in COMMISSION_RATES['30%']:
        rate = 0.30
        rate_group = '30%'
    elif account in COMMISSION_RATES['50%']:
        rate = 0.50
        rate_group = '50%'
    else:
        rate = 0.35
        rate_group = '35%'

    # 确定管理组
    management_group = None
    for group_name, group_accounts in ACCOUNT_GROUPS.items():
        if account in group_accounts:
            management_group = group_name
            break

    # 计算每个股票的收益和分成
    stock_details = []
    total_revenue = 0
    total_commission = 0
    total_loss = 0

    for stock in stocks:
        if 'shouyi_col' not in stock:
            continue

        shouyi_col = stock['shouyi_col']
        value = ws.cell(row=row_idx, column=shouyi_col + 1).value

        if value is not None and isinstance(value, (int, float)):
            revenue = round(value, 2)
            total_revenue += revenue

            # 检查是否有特殊分成规则
            stock_name_short = stock['name'].split('(')[0]  # 去掉日期部分
            special_key = (account, stock_name_short)
            is_special = False
            special_note = None

            if special_key in SPECIAL_COMMISSIONS:
                # 特殊分成规则
                if revenue > 0:
                    commission = SPECIAL_COMMISSIONS[special_key]
                    total_commission += commission
                    is_special = True
                    special_note = '乙组资金固定分配'
                else:
                    commission = 0
                    total_loss += abs(revenue)
            else:
                # 正常分成规则
                if revenue > 0:
                    commission = round(revenue * rate, 2)
                    total_commission += commission
                else:
                    commission = 0
                    total_loss += abs(revenue)

            stock_details.append({
                'stock_name': stock['name'],
                'revenue': revenue,
                'commission': commission,
                'is_special': is_special,
                'special_note': special_note
            })

    account_details.append({
        'account': account,
        'rate': rate,
        'rate_group': rate_group,
        'management_group': management_group,
        'total_revenue': round(total_revenue, 2),
        'total_commission': round(total_commission, 2),
        'total_loss': round(total_loss, 2),
        'stocks': stock_details
    })

print(f"   计算完成")

# 计算特定范围分成（长风药业→壁仞科技）
print("[6/6] 计算特定范围分成...")

# 找到长风药业、壁仞科技、紫金国际的索引
changfeng_idx = next((i for i, s in enumerate(stocks) if '长风药业' in s['name']), None)
biren_idx = next((i for i, s in enumerate(stocks) if '壁仞科技' in s['name']), None)
zijin_idx = next((i for i, s in enumerate(stocks) if '紫金国际' in s['name']), None)

special_range_data = []

for account in SPECIAL_RANGE_ACCOUNTS:
    if account not in account_rows:
        continue

    # 找到该账户的详情
    acc_detail = next((a for a in account_details if a['account'] == account), None)
    if not acc_detail:
        continue

    # 计算范围内的分成
    range_commission = 0
    range_revenue = 0
    range_stock_details = []  # 保存范围内的股票详情

    # 长风药业到壁仞科技
    for stock_detail in acc_detail['stocks']:
        stock_name = stock_detail['stock_name']
        stock_idx = next((i for i, s in enumerate(stocks) if s['name'] == stock_name), None)

        if stock_idx is not None and changfeng_idx <= stock_idx <= biren_idx:
            range_commission += stock_detail['commission']
            range_revenue += stock_detail['revenue']
            range_stock_details.append(stock_detail)

    # 特殊账户额外加紫金国际
    if account in ZIJIN_EXTRA_ACCOUNTS and zijin_idx is not None:
        zijin_stock = stocks[zijin_idx]
        zijin_detail = next((s for s in acc_detail['stocks'] if s['stock_name'] == zijin_stock['name']), None)
        if zijin_detail:
            range_commission += zijin_detail['commission']
            range_revenue += zijin_detail['revenue']
            range_stock_details.append(zijin_detail)

    special_range_data.append({
        'account': account,
        'rate_group': acc_detail['rate_group'],
        'management_group': acc_detail['management_group'],
        'range_revenue': round(range_revenue, 2),
        'range_commission': round(range_commission, 2),
        'has_zijin_extra': account in ZIJIN_EXTRA_ACCOUNTS,
        'stocks': range_stock_details  # 添加股票详情
    })

print(f"   特定范围分成计算完成，涉及 {len(special_range_data)} 个账户")

# 查找缺失信息
print("\n查找缺失信息...")
zhongqian_cols = []
maichu_cols = []

for i in range(len(row2)):
    col_title = str(row2[i].value) if row2[i].value else ""
    if "中签情况" in col_title:
        zhongqian_cols.append(i)
    elif "卖出价格" in col_title:
        maichu_cols.append(i)

# 创建中签情况到卖出价格的映射
zhongqian_to_maichu = {}
for zq_col in zhongqian_cols:
    stock_name = None
    for j in range(zq_col, -1, -1):
        if row1[j].value:
            stock_name = row1[j].value
            break
    for mc_col in maichu_cols:
        mc_stock_name = None
        for j in range(mc_col, -1, -1):
            if row1[j].value:
                mc_stock_name = row1[j].value
                break
        if stock_name == mc_stock_name:
            zhongqian_to_maichu[zq_col] = mc_col
            break

missing_price_records = []
for account in accounts:
    row_idx = account_rows[account]
    for zq_col, mc_col in zhongqian_to_maichu.items():
        zq_value = str(ws.cell(row=row_idx, column=zq_col + 1).value or "")
        mc_value = ws.cell(row=row_idx, column=mc_col + 1).value

        if "已中签" in zq_value and (mc_value is None or str(mc_value).strip() == ""):
            stock_name = None
            for j in range(zq_col, -1, -1):
                if row1[j].value:
                    stock_name = row1[j].value
                    break

            missing_price_records.append({
                'account': account,
                'stock': stock_name,
                'row': row_idx,
                'col': get_column_letter(mc_col + 1)
            })

print(f"   找到 {len(missing_price_records)} 条缺失记录")

# 生成JSON数据
print("\n生成JSON数据...")
output_data = {
    'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    'summary': {
        'total_revenue': round(sum(s['revenue'] for s in stock_revenues), 2),
        'total_stocks': len(stocks),
        'total_accounts': len(accounts),
        'total_commission': round(sum(a['total_commission'] for a in account_details), 2),
        'total_loss': round(sum(a['total_loss'] for a in account_details), 2)
    },
    'stocks': stock_revenues,
    'accounts': account_details,
    'special_range': special_range_data,
    'missing_records': missing_price_records,
    'commission_rates': {
        '0%': COMMISSION_RATES['0%'],
        '30%': COMMISSION_RATES['30%'],
        '50%': COMMISSION_RATES['50%'],
        '35%': COMMISSION_RATES['35%']
    },
    'account_groups': ACCOUNT_GROUPS,
    'special_commissions': {f"{k[0]}_{k[1]}": v for k, v in SPECIAL_COMMISSIONS.items()}
}

# 保存JSON文件
output_file = '/Users/wright/Documents/WFCode/GangGu/report_data.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print(f"\n✓ 数据已保存到: {output_file}")
print("\n" + "=" * 100)
print("数据生成完成！")
print("=" * 100)
print(f"\n统计摘要:")
print(f"  总收益: {output_data['summary']['total_revenue']:,.2f} 元")
print(f"  总分成: {output_data['summary']['total_commission']:,.2f} 元")
print(f"  总亏损: {output_data['summary']['total_loss']:,.2f} 元")
print(f"  股票数: {output_data['summary']['total_stocks']}")
print(f"  账户数: {output_data['summary']['total_accounts']}")
print(f"  缺失记录: {len(missing_price_records)} 条")
